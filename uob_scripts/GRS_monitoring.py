#       
#	GRS_monitoring.py
#
#	The purpose of this script is to notify relevant personel about a student
#	and supervisors completion of monthly GRS forms during post-graduate (PG)
#	studies. This is based on the monthly assignments made on Canvas for the
#	student to upload their completed GRS form to.
#
#	This script reads an Excel file of four columns:
#		student_name, student_login_id, supervisor_name, supervisor_login_id
#	and loops through every student on that list. While looping through every
#	student in the Excel file, it checks whether a GRS submission is present
#	for the student and whether it has been marked as complete by the
#	supervisor. Depending on this, and the working day of the month, the
#	script sends emails to the student, supervisor, etc.
#
#	Things that need to be set:
#
#	course_id - the Canvas course on which GRS assignments are put
course_id = '7054'
#
#	form_submission_reminder_dates - dates for form submission reminders
form_submission_reminder_dates = [
	[6, 14], # dates when to email student only
	[16, 18], # dates when to email student and supervisor
	[19], # first escalation emails
	[30, 32] # second escalation emails
]
#form_submission_reminder_dates = [
#	[1, 3],
#	[5, 7, 9, 11],
#	[13],
#	[15]
#]
#
#	form_submission_reminder_dates - dates for form submission reminders
mark_complete_reminder_dates = [
	[5, 10, 15, 18], # dates when to email supervisor only
	[19], # first escalation emails
	[30, 32] # second escalation emails
]
#mark_complete_reminder_dates = [
#	[1, 3, 5],
#	[7, 9, 11, 13],
#	[15]
#]
#
#	summary_email_dates - dates for summary email
summary_email_dates = [14, 15, 16, 17, 18, 20]
#
#	first_escalation_emails - set of emails for first escalation
first_escalation_emails = ['a.mottura@bham.ac.uk']
#
#	second_escalation_emails - set of emails for second escalation
second_escalation_emails = ['i.hoffman@bham.ac.uk']
#
#	PG office email - set the email of PG office
PG_office_email = ['metmat-phdmscadmin@contacts.bham.ac.uk']
#
#	student_list - path to Excel file containing details of students
student_list = '/mnt/metadmin/CANVASBOTS/PG/GRS_Monitoring/Student_List.xlsx'
#
#	NOTE: this script sends email using the class defined in
#	PyCAPI/uob_scripts/uob_utils.py
#	If you wish to use the script, you need to check that you set up email
#	appropriately.
#
#
#
import PyCAPI
import uob_utils
from openpyxl import load_workbook
import datetime
import calendar


capi = PyCAPI.CanvasAPI()
mail = uob_utils.MailAPI()
todaydate = datetime.date.today()








###############################################################################
# Calculate the current working day
#
firstdate = todaydate.replace(day=1)
workingday = ((todaydate.day - max(5 - firstdate.weekday(),0) - min(todaydate.weekday()+1,5))/7)*5 + max(5 - firstdate.weekday(),0) + min(todaydate.weekday()+1,5)




###############################################################################
# Obtain student details from Excel document
#
wb = load_workbook(student_list, read_only = True, data_only = True)
ws = wb['Sheet1']
students = []
i = 2
while ws['A'+str(i)].value != None:
	students.append({'name':ws['A'+str(i)].value, 'id':ws['B'+str(i)].value, 'supervisor_name':ws['C'+str(i)].value, 'supervisor_id':ws['D'+str(i)].value, 'skip': False})
	if ws['E'+str(i)].value != None:
		students[-1]['skip'] = True
	i += 1   




###############################################################################
# Obtain further student details from Canvas
#
for student in students:
	extra_details = capi.get_user('sis_login_id:'+str(student['id']))
	student['canvas_id'] = extra_details['id']
	student['email'] = extra_details['primary_email']
	student['supervisor_email'] = capi.get_user('sis_login_id:'+str(student['supervisor_id']))['primary_email']


###############################################################################
# Enroll all students and supervisors not already enrolled on GRS course
#
enrollments = capi.get('/courses/%s/enrollments' % course_id)
enrolled = []
for enrollment in enrollments:
	if 'sis_login_id' in enrollment['user']:
		enrolled.append(enrollment['user']['sis_login_id'])
for student in students:
	if not student['id'] in enrolled:
		payload = {}
		payload['enrollment[user_id]'] = student['canvas_id']
		payload['enrollment[type]'] = 'StudentEnrollment'
		payload['enrollment[enrollment_state]'] = 'active'
		capi.post('/courses/%s/enrollments' % course_id, payload=payload)
	if not student['supervisor_id'] in enrolled:
		payload = {}
		payload['enrollment[user_id]'] = 'sis_login_id:'+str(student['supervisor_id'])
		payload['enrollment[type]'] = 'TeacherEnrollment'
		payload['enrollment[enrollment_state]'] = 'active'
		capi.post('/courses/%s/enrollments' % course_id, payload=payload)




###############################################################################
# Get completion details of the current months GRS assignment
#
# Using deadlines to get current assignment
assignments = capi.get_assignments(course_id)
for assignment in assignments:
	due_datetime = datetime.datetime.strptime(assignment['due_at'], '%Y-%m-%dT%H:%M:%SZ')
	if todaydate.month == due_datetime.month and todaydate.year == due_datetime.year:
		assignment_id = assignment['id']
	next_assignment = False
	if todaydate.month == 12:
		if due_datetime.month == 1 and due_datetime.year == todaydate.year+1:
			next_assignment = True
	elif todaydate.month+1 == due_datetime.month and due_datetime.year == todaydate.year:
		next_assignment = True


for student in students:
	# Download details of submission for the user
	submission = capi.get_assignment_submissions(course_id, assignment_ids=assignment_id, user_ids=student['canvas_id'], grouped=False)
	# Check whether form has been uploaded
	if 'attachments' in submission[0]['submissions'][0]:
		student['form'] = True
	else:
		student['form'] = False
	# Check whether submission has been marked as complete
	if 'grade' in submission[0]['submissions'][0]:
		if submission[0]['submissions'][0]['grade'] == 'complete':
			student['complete'] = True
		else:
			student['complete'] = False
	else:
		student['complete'] = False




###############################################################################
# Prepare emails based on completion and workinday
#
for student in students:
	recipients = []
	cc_recipients = []
	message_subject = ''
	message_body = ''
	
	# If student is to be skipped, then continue
	if student['skip'] == True:
		continue
	
	# If it is marked as complete, but file is not uploaded, mark as incomplete
	if student['form'] == False and student['complete'] == True:
		print 'Mark as incomplete'
		student['complete'] = False
		# Functionality needed that will set assignment to uncomplete - NEEDS IMPLEMENTATION

		
	# If file has not been uploaded and complete mark is missing
	if student['form'] == False and student['complete'] == False:
		if workingday in form_submission_reminder_dates[0]: # just a reminder to the student
			message_subject = 'Reminder: your GRS2 form'
			message_body = """Hi %s,
This is a reminder that you should have your monthly official supervision meeting with your supervisor, %s, as soon as possible.
Contact your supervisor and plan a suitable time within the next few days.
Make sure you complete part A of the GRS2 form ahead of the meeting.
You can download the GRS2 form from Canvas, at https://canvas.bham.ac.uk/courses/7054 
If you have already had the monthly official supervision meeting for this month, please upload the fully completed GRS2 form as soon as possible.
If you are currently on leave and unable to meet with your supervisor, please notify the PG office by replying to this email.
Best Regards,
Met&Mat PG Office
			""" % (student['name'], student['supervisor_name'])
			recipients.append(student['email'])
		elif workingday in form_submission_reminder_dates[1]: # just a reminder to the student, but includes the supervisor
			message_subject = 'REMINDER: your GRS2 form'
			message_body = """Hi %s,
This is an IMPORTANT reminder that you should have your monthly official supervision meeting with your supervisor, %s, as soon as possible.
Contact your supervisor and plan a suitable time within the next few days.
Your supervisor has been CC'd to this email as well, so he may get in touch with you directly.
Make sure you complete part A of the GRS2 form ahead of the meeting.
You can download the GRS2 form from Canvas, at https://canvas.bham.ac.uk/courses/7054 
If you have already had the monthly official supervision meeting for this month, please upload the fully completed GRS2 form as soon as possible.
If you are currently on leave and unable to meet with your supervisor, please notify the PG office by replying to this email.
Best Regards,
Met&Mat PG Office
			""" % (student['name'], student['supervisor_name'])
			recipients.append(student['email'])
			cc_recipients.append(student['supervisor_email'])
		elif workingday in form_submission_reminder_dates[2]: # first escalation
			message_subject = 'IMPORTANT: your GRS2 form'
			message_body = """Hi %s,
It is VERY IMPORTANT that you have your monthly official supervision meeting with your supervisor, %s, as soon as possible!
Contact your supervisor and plan a suitable time today, or at the earliest possible time.
Make sure you complete part A of the GRS2 form ahead of the meeting.
You can download the GRS2 form from Canvas, at https://canvas.bham.ac.uk/courses/7054 
If you have already had the monthly official supervision meeting for this month, please upload the fully completed GRS2 form now.
If you are currently on leave and unable to meet with your supervisor, please notify the PG office by replying to this email.
Best Regards,
Met&Mat PG Office
			""" % (student['name'], student['supervisor_name'])
			recipients.append(student['email'])
			cc_recipients.append(student['supervisor_email'])
			cc_recipients.extend(first_escalation_emails)
		elif workingday in form_submission_reminder_dates[3]: # second and final escalation
			message_subject = 'ATTENTION REQUIRED: your GRS2 form'
			message_body = """Hi %s and %s,
The last working day of the month is approaching quickly, and an official supervision meeting MUST happen soon!
Please complete part A of the GRS2 form, meet with your supervisor to complete the rest, and upload the fully completed form to Canvas.
You are strongly encouraged to do this TODAY!
You can download the GRS2 form from Canvas, at https://canvas.bham.ac.uk/courses/7054 
If you have already had the monthly official supervision meeting for this month, please upload the fully completed GRS2 form now.
If you are currently on leave and unable to meet with your supervisor, please notify the PG office by replying to this email.
Best Regards,
Met&Mat PG Office
			""" % (student['name'], student['supervisor_name'])
			recipients.append(student['email'])
			cc_recipients.append(student['supervisor_email'])
			cc_recipients.extend(first_escalation_emails)
			cc_recipients.extend(second_escalation_emails)
			
			
			
	# If file has been uploaded but complete mark is missing
	if student['form'] == True and student['complete'] == False:
		if workingday in mark_complete_reminder_dates[0]: # just a reminder to the supervisor
			message_subject = 'Reminder: mark GRS2 form as complete'
			message_body = """Hi %s,
Your student, %s, has uploaded a GRS2 form to Canvas.
Please check the uploaded form is complete and is an accurate representation of your official monthly supervision meeting, then mark it as complete.
You can do so by visiting https://canvas.bham.ac.uk/courses/7054/gradebook/speed_grader?assignment_id=%s#%%7B%%22student_id%%22%%3A%%22%s%%22%%7D
If you are currently on leave and unable to do this, please notify the PG office by replying to this email.
Best Regards,
Met&Mat PG Office
			""" % (student['supervisor_name'],student['name'],assignment_id,student['canvas_id'])
			recipients.append(student['supervisor_email'])
		elif workingday in mark_complete_reminder_dates[1]: # first escalation
			message_subject = 'REMINDER: mark GRS2 form as complete'
			message_body = """Hi %s,
Your student, %s, has uploaded a GRS2 form to Canvas, and this needs to be marked as complete as soon as possible.
Please check the uploaded form is complete and is an accurate representation of your official monthly supervision meeting, then mark it as complete.
You can do so by visiting https://canvas.bham.ac.uk/courses/7054/gradebook/speed_grader?assignment_id=%s#%%7B%%22student_id%%22%%3A%%22%s%%22%%7D
If you are currently on leave and unable to do this, please notify the PG office by replying to this email.
Note that this needs to be done quickly, as the end of the month is approaching.
Best Regards,
Met&Mat PG Office
			""" % (student['supervisor_name'],student['name'],assignment_id,student['canvas_id'])
			recipients.append(student['supervisor_email'])
			cc_recipients.extend(first_escalation_emails)
		elif workingday in mark_complete_reminder_dates[2]: # second escalation
			message_subject = 'ATTENTION REQUIRED: mark GRS2 form as complete'
			message_body = """Hi %s,
Your student, %s, has uploaded a GRS2 form to Canvas, and this needs to be marked as complete now!
Please check the uploaded form is complete and is an accurate representation of your official monthly supervision meeting, then mark it as complete.
You can do so by visiting https://canvas.bham.ac.uk/courses/7054/gradebook/speed_grader?assignment_id=%s#%%7B%%22student_id%%22%%3A%%22%s%%22%%7D
If you are currently on leave and unable to do this, please notify the PG office by replying to this email.
Please do this now!
Best Regards,
Met&Mat PG Office
			""" % (student['supervisor_name'],student['name'],assignment_id,student['canvas_id'])
			recipients.append(student['supervisor_email'])
			cc_recipients.extend(first_escalation_emails)	
			cc_recipients.extend(second_escalation_emails)


	if len(cc_recipients) == 0:
		msg = uob_utils.EMailMessage(", ".join(recipients), message_subject)
	else:
		msg = uob_utils.EMailMessage(", ".join(recipients), message_subject, cc_addr=", ".join(cc_recipients))
	msg.body(message_body)
	all_recipients = recipients + cc_recipients
	if len(all_recipients) != 0:
		mail.send(all_recipients, msg) # send email




###############################################################################
# Send summary email
#
if workingday in summary_email_dates:
	message_subject = 'GRS2 Form Summary'
	message_body = """Hi,
Here is the summary of GRS2 form submissions.
	"""
	
	message_body += '\nThese students have not submitted a form yet:\n'
	for student in students:
		if student['form'] == False and student['complete'] == False:
			message_body += '  - ' + student['name'].encode('ascii') + ' \n' 
	
	message_body += '\nThese students have submitted a form, but it still needs to be marked as complete by the supervisor:\n'
	for student in students:
		if student['form'] == True and student['complete'] == False:
			message_body += '  - ' + student['name'].encode('ascii') + ' \n'
	
	message_body += '\nThese students have completed the GRS2 form requirements:\n'
	for student in students:
		if student['form'] == True and student['complete'] == True:
			message_body += '  - ' + student['name'].encode('ascii') + ' \n'
	
	message_body += '\n'
	message_body += 'Bye bye,\nGRS2 form monitoring'
	msg = uob_utils.EMailMessage(", ".join(PG_office_email), message_subject)
	msg.body(message_body)
	mail.send(PG_office_email, msg)








###############################################################################
# Creating an assignment for the next month if there is currently none
#
# List of assignment properties that need to be copied:
keylist = [
	'submission_types',
	'allowed_extensions',
	'turnitin_enabled',
	'vericite_enabled',
	'vericite_enabled',
	'integration_data',
	'integration_id',
	'notify_of_update',
	'group_category_id',
	'grade_group_students_individually',
	'external_tool_tag_attributes',
	'automatic_peer_reviews',
	'points_possible',
	'grading_type',
	'muted',
	'assignment_overrides',
	'only_visible_to_overrides',
	'published',
	'grading_standard_id',
	'omit_from_final_grade',
	'quiz_lti',
	'assignment_group_id',
	'description'
]
# Keys to be copied that contain a list:
listKeys = [
	'submission_types',
	'allowed_extensions'
]
# Check whether an assignment is present for the following month
if workingday > 18 and next_assignment == False: # New assignment needed for next month
	current_assignment = capi.get_assignment(course_id, assignment_id)
	payload = {} # Create payload which contains new assignment information
	for key in current_assignment: # Copy keys from old assignment to new assignment
		if key in keylist:
			if key in listKeys:
				payload['assignment['+key+'][]'] = current_assignment[key]
			else:
				payload['assignment['+key+']'] = current_assignment[key]
	# Figure out month and year of next assignment
	if datetime.datetime.strptime(current_assignment['due_at'], '%Y-%m-%dT%H:%M:%SZ').month == 12:
		nextmonthnum = 1
		nextmonthyear = datetime.datetime.strptime(current_assignment['due_at'], '%Y-%m-%dT%H:%M:%SZ').year+1
	else:
		nextmonthnum = datetime.datetime.strptime(current_assignment['due_at'], '%Y-%m-%dT%H:%M:%SZ').month+1
		nextmonthyear = datetime.datetime.strptime(current_assignment['due_at'], '%Y-%m-%dT%H:%M:%SZ').year
	# Create name and set other settings for new assignment
	payload['assignment[name]'] = 'GRS2 Form - ' + calendar.month_name[nextmonthnum] + ' ' + str(nextmonthyear)
	payload['assignment[position]'] = current_assignment['position']+1
	payload['assignment[peer_reviews]'] = 'false'
	payload['assignment[turnitin_enabled]'] = 'false'
	payload['assignment[omit_from_final_grade]'] = 'false'
	unlock_at = datetime.datetime.strptime(current_assignment['due_at'], '%Y-%m-%dT%H:%M:%SZ').replace(year=nextmonthyear).replace(month=nextmonthnum).replace(day=1)
	payload['assignment[unlock_at]'] = PyCAPI.datetime2unicode(unlock_at)
	due_at = datetime.datetime.strptime(current_assignment['due_at'], '%Y-%m-%dT%H:%M:%SZ').replace(year=nextmonthyear).replace(month=nextmonthnum).replace(day=27)
	payload['assignment[due_at]'] = PyCAPI.datetime2unicode(due_at)
	lock_at = datetime.datetime.strptime(current_assignment['due_at'], '%Y-%m-%dT%H:%M:%SZ').replace(year=nextmonthyear).replace(month=nextmonthnum).replace(day=28)
	payload['assignment[lock_at]'] = PyCAPI.datetime2unicode(lock_at)
	capi.post('/courses/%s/assignments' % course_id, payload=payload)



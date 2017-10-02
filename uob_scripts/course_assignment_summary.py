#
#       course_assignment_summary.py
#
#       This code will create an Excel workbook using openpyxl which will
#       contain summary details of courses and assignments for the users account
#
#       This script will also send feedback reminders to relevant memebers of
#       staff if assignments are left ungraded after a set number of working days
#       from the assignment due date
#
#       Things that need to be set:
#
#       output_dir - path of Excel file to be saved
output_dir = '/mnt/metadmin/CANVASBOTS/UG/Courses_and_Assignments/'
#
#       name_extenstion - name of Excel file after Canvas course number
name_extension = 'assignment_summary'
#
#
#       included_terms - list of terms that the user wishes to recieve when
#       running the code, format = 'yyyy/yy' or 'Default term'
included_terms = ['2017/18']
#
#
#       offset_days - working days after which the due date of an assignment
#       should be
offset_days = 14
#
#
#       reminder_dates - days from due date for reminder emails to be sent
reminder_dates = [
    [14], # day to email teachers
    [20] # day to email teachers and TSO
    ]
#
#
#       TSO_email - email of TSO
TSO_email = ['me@mottura.org']
#
#
#
import sys
sys.path.append("../module/") # First two lines are needed for import of PyCAPI
import PyCAPI
import uob_utils
import datetime
import json
import numpy as np
import pandas
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill
from openpyxl import worksheet


capi = PyCAPI.CanvasAPI()
mail = uob_utils.MailAPI()
today = datetime.datetime.now()



###############################################################################
# Create workbook for storing information in
#
wb = Workbook()
wb.remove(wb.active) # Remove initially created sheet
ws = wb.create_sheet(title='Courses') # Set active worksheet and name



###############################################################################
# Set columns for course summary sheet (this way one needs to do this once)
#
col_term = 'A'
col_cid = 'B'
col_cnm = 'C'
col_ccd = 'D'
col_cadmin = 'E'
col_cavail = 'F'
col_tenroll = 'G'
col_senroll = 'H'

# Format course summary sheet
ws[col_term+'1'] = 'Term'
ws[col_cid+'1'] = 'Course ID'
ws[col_cnm+'1'] = 'Course Name'
ws[col_ccd+'1'] = 'Course Code'
ws[col_cadmin+'1'] = 'Admin'
ws[col_cavail+'1'] = 'Availability'
ws[col_tenroll+'1'] = 'Teachers Enrolled'
ws[col_senroll+'1'] = 'Students Enrolled'

# Set column widths if needed
ws.column_dimensions[col_term].width = 10
ws.column_dimensions[col_cid].width = 11
ws.column_dimensions[col_cnm].width = 30
ws.column_dimensions[col_ccd].width = 15
ws.column_dimensions[col_cadmin].width = 15
ws.column_dimensions[col_cavail].width = 15
ws.column_dimensions[col_tenroll].width = 20
ws.column_dimensions[col_senroll].width = 20



###############################################################################
# Retrieve data of all courses from Canvas account
#
# Following 4 lines are needed to return courses that only I am a teacher on
#payload = {}
#payload['enrollment_type'] = 'teacher' # Remove this line if used by admin
#payload['include[]'] = ['term', 'teachers', 'total_students']
#courses = capi.get('/courses', payload=payload) # Remove this line if used by admin
#courses = capi.get_courses(payload=payload) # Use this line when used by Met&Mat admin
#print json.dumps(courses, indent = 2)
#allcourses = capi.get('/courses', payload=payload)
allcourses = capi.get_courses(account_id='114', include=['term', 'teachers', 'total_students'])
courses = []
for course in allcourses:
    if course['term']['name'] in included_terms:
        courses.append(course)
#print json.dumps(courses, indent = 2)


i = 2
for course in courses: # Loop through all courses in Canvas account
    ws[col_term+str(i)] = course['term']['name']
    ws[col_cid+str(i)] = '=HYPERLINK("http://147.188.152.33:8080/eval.php?url=https://canvas.bham.ac.uk/courses/'+str(course['id'])+'", "'+str(course['id'])+'")'
    ws[col_cnm+str(i)] = '=HYPERLINK("http://147.188.152.33:8080/eval.php?url=https://canvas.bham.ac.uk/courses/'+str(course['id'])+'", "'+str(course['name'])+'")'
    ws[col_ccd+str(i)] = '=HYPERLINK("http://147.188.152.33:8080/eval.php?url=https://canvas.bham.ac.uk/courses/'+str(course['id'])+'", "'+str(course['course_code'])+'")'
    ws[col_cadmin+str(i)] = course['account_id']
    ws[col_cavail+str(i)] = course['workflow_state'].title()
    if course['workflow_state'] != 'available':
        for cell in ws[str(i)+':'+str(i)]:
            cell.font = Font(color='00D3D3D3')
    ws[col_tenroll+str(i)] = len(course['teachers'])
    ws[col_senroll+str(i)] = course['total_students']
    i += 1

# Make the workbook into a filtered table.
ws.auto_filter.ref = 'A1:AA'+str(i)



###############################################################################
# Summarise assignments in a different Excel sheet
#
ws = wb.create_sheet(title='Assignments') # Set active worksheet



###############################################################################
# Set columns for assignment summary sheet (this way one needs to do this once)
#
col_cid = 'A'
col_cnm = 'B'
col_ccd = 'C'
col_term = 'D'
col_asgnid = 'E'
col_asgnnm = 'F'
col_cadmin = 'G'
col_cavail = 'H'
col_asgnavail = 'I'
col_asgntype = 'J'
col_group = 'K'
col_unlock = 'L'
col_due = 'M'
col_lock = 'N'
col_sub = 'O'
col_ungr = 'P'
col_miss = 'Q'
col_late = 'R'
col_median = 'S'
col_grby = 'T'
col_fingr = 'U'
col_weight = 'V'
col_muted = 'W'


# Format assignment summary sheet
ws[col_cid+'1'] = 'Course ID'
ws[col_cnm+'1'] = 'Course Name'
ws[col_ccd+'1'] = 'Course Code'
ws[col_term+'1'] = 'Term'
ws[col_asgnid+'1'] = 'Assignment ID'
ws[col_asgnnm+'1'] = 'Assignment Name'
ws[col_cadmin+'1'] = 'Admin'
ws[col_cavail+'1'] = 'Course Availability'
ws[col_asgnavail+'1'] = 'Assignment Availability'
ws[col_asgntype+'1'] = 'Assignment Type'
ws[col_group+'1'] = 'Grouped'
ws[col_unlock+'1'] = 'Unlock at'
ws[col_due+'1'] = 'Due at'
ws[col_lock+'1'] = 'Lock at'
ws[col_sub+'1'] = 'Submissions'
ws[col_ungr+'1'] = 'Ungraded'
ws[col_miss+'1'] = 'Missing'
ws[col_late+'1'] = 'Late'
ws[col_median+'1'] = 'Median'
ws[col_grby+'1'] = 'Grade by'
ws[col_fingr+'1'] = 'Final grade'
ws[col_weight+'1'] = 'Weighting'
ws[col_muted+'1'] = 'Muted'

# Set column widths if needed
ws.column_dimensions[col_cid].width = 11
ws.column_dimensions[col_cnm].width = 20
ws.column_dimensions[col_ccd].width = 15
ws.column_dimensions[col_asgnid].width = 14
ws.column_dimensions[col_asgnnm].width = 30
ws.column_dimensions[col_cadmin].width = 15
ws.column_dimensions[col_cavail].width = 17
ws.column_dimensions[col_asgnavail].width = 17
ws.column_dimensions[col_asgntype].width = 16
ws.column_dimensions[col_unlock].width = 17
ws.column_dimensions[col_due].width = 17
ws.column_dimensions[col_lock].width = 17
ws.column_dimensions[col_sub].width = 12
ws.column_dimensions[col_ungr].width = 12
ws.column_dimensions[col_grby].width = 12
ws.column_dimensions[col_fingr].width = 12
ws.column_dimensions[col_weight].width = 12



###############################################################################
# Retrieve data of all assignments from Canvas account
#
i = 2
for course in courses:
    course_assignments = capi.get_assignments(course['id'])
    payload = {}
    payload['include[]'] = ['assignments']
    assignment_groups = capi.get('/courses/%s/assignment_groups' % course['id'], payload=payload)
    analytics_missing = False
    try:
        analytics = capi.get('/courses/%s/analytics/assignments' % course['id'])
    except:
        analytics_missing = True
    #print json.dumps(course_assignments, indent = 2)
    #print json.dumps(assignment_groups, indent = 2)
    for assignment in course_assignments:
        ws[col_cid+str(i)] = '=HYPERLINK("http://147.188.152.33:8080/eval.php?url=https://canvas.bham.ac.uk/courses/'+str(course['id'])+'", "'+str(course['id'])+'")'
        ws[col_cnm+str(i)] = '=HYPERLINK("http://147.188.152.33:8080/eval.php?url=https://canvas.bham.ac.uk/courses/'+str(course['id'])+'", "'+str(course['name'])+'")'
        ws[col_ccd+str(i)] = '=HYPERLINK("http://147.188.152.33:8080/eval.php?url=https://canvas.bham.ac.uk/courses/'+str(course['id'])+'", "'+str(course['course_code'])+'")'  
        ws[col_term+str(i)] = course['term']['name']
        ws[col_asgnid+str(i)] = '=HYPERLINK("http://147.188.152.33:8080/eval.php?url=https://canvas.bham.ac.uk/courses/'+str(course['id'])+'/assignments/'+str(assignment['id'])+'", "'+str(assignment['id'])+'")'
        ws[col_asgnnm+str(i)] = '=HYPERLINK("http://147.188.152.33:8080/eval.php?url=https://canvas.bham.ac.uk/courses/'+str(course['id'])+'/assignments/'+str(assignment['id'])+'", "'+str(assignment['name'].encode('ascii', 'ignore'))+'")'
        ws[col_cadmin+str(i)] = course['account_id']
        ws[col_cavail+str(i)]= course['workflow_state'].title()
        if assignment['published']:
            ws[col_asgnavail+str(i)] = 'Published'
        else:
            ws[col_asgnavail+str(i)] = 'Unpublished'
        if 'online_upload' in assignment['submission_types'] or 'online_quiz' in assignment['submission_types']:
            ws[col_asgntype+str(i)] = 'Online'
        else:
            ws[col_asgntype+str(i)] = 'Paper'
        if assignment['group_category_id'] == None:
            ws[col_group+str(i)] = 'No'
        else:
            ws[col_group+str(i)] = 'Yes'
        if assignment['unlock_at'] == None:
            ws[col_unlock+str(i)] = 'None set'
        else:
            ws[col_unlock+str(i)] = str(datetime.datetime.strptime(assignment['unlock_at'], "%Y-%m-%dT%H:%M:%SZ"))
        if assignment['due_at'] == None:
            ws[col_due+str(i)] = 'None set'
            ws[col_grby+str(i)] = 'None set'
        else:
            due_date = datetime.datetime.strptime(assignment['due_at'], "%Y-%m-%dT%H:%M:%SZ")
            ws[col_due+str(i)] = str(due_date)
            grade_by = ((pandas.Timestamp(np.busday_offset(due_date, offset_days, roll='forward'))).to_pydatetime()).date()
            ws[col_grby+str(i)] = str(grade_by)
            recipients = []
            cc_recipients = []
            if (int(np.busday_count(due_date,today)) in reminder_dates[0]) and assignment['needs_grading_count'] != 0: # If reminder email for teacher is needed
                # Collects email of all teachers for assignment to email
                for course_teacher in course['teachers']:
                    teacher = capi.get_user(course_teacher['id'])
                    recipients.append(teacher['primary_email'])
                message_subject = '*CANVAS TESTING* IMPORTANT: There are assignments ungraded'
                message_body = """Hi,

This is an IMPORTANT reminder that the due date for grading the assignment: %s has been reached and there are still submissions left ungraded.
Please can this be addressed as soon as possible.
Link: %s

Best Regards,
Met&Mat Office
			""" % (assignment['name'], assignment['html_url'])
                # Now send email to teachers
                msg = uob_utils.EMailMessage(", ".join(recipients), message_subject)
                msg.body(message_body)
                all_recipients = recipients + cc_recipients
                #mail.send(all_recipients, msg) # Send email
            elif (int(np.busday_count(due_date,today)) in reminder_dates[1]) and assignment['needs_grading_count'] != 0: # If reminder email for teacher and TSO is needed
                # Collects email of all teachers for assignment to email
                for course_teacher in course['teachers']:
                    teacher = capi.get_user(course_teacher['id'])
                    recipients.append(teacher['primary_email'])
                cc_recipients.extend(TSO_email)
                message_subject = '*CANVAS TESTING* ATTENTION REQUIRED: Assignment submissions still ungraded'
                message_body = """Hi,

This is an VERY IMPORTANT reminder that the due date for grading the assignment: %s has been passed and there are still submissions left ungraded.
This must be addressed immediately.
Link: %s

Best Regards,
Met&Mat Office
			""" % (assignment['name'], assignment['html_url'])
                # Now send email to teachers and TSO
                msg = uob_utils.EMailMessage(", ".join(recipients), message_subject, cc_addr=", ".join(cc_recipients))
                msg.body(message_body)
                all_recipients = recipients + cc_recipients
                #mail.send(all_recipients, msg) # Send email
        if assignment['lock_at'] == None:
            ws[col_lock+str(i)] = 'None set'
        else:
            ws[col_lock+str(i)] = str(datetime.datetime.strptime(assignment['lock_at'], "%Y-%m-%dT%H:%M:%SZ"))
        #ws[col_sub+str(i)] =
        ws[col_ungr+str(i)] = assignment['needs_grading_count']
        if analytics_missing == False:
            for analytic in analytics:
                if analytic['assignment_id'] == assignment['id']:
                    if analytic['tardiness_breakdown']['missing'] == None:
                        ws[col_miss+str(i)] = 'None'
                    else:
                        ws[col_miss+str(i)] = analytic['tardiness_breakdown']['missing']
                    if analytic['tardiness_breakdown']['late'] == None:
                        ws[col_late+str(i)] = 'None'
                    else:
                        ws[col_late+str(i)] = analytic['tardiness_breakdown']['late']
                    if analytic['median'] == None:
                        ws[col_median+str(i)] = 'None'
                    else:
                        ws[col_median+str(i)] = analytic['median']
        else:
            ws[col_miss+str(i)] = 'Missing'
            ws[col_late+str(i)] = 'Missing'
            ws[col_median+str(i)] = 'Missing'
        if assignment['omit_from_final_grade'] == False:
            ws[col_fingr+str(i)] = 'Yes'
        else:
            ws[col_fingr+str(i)] = 'No'
        for assignment_group in assignment_groups:
            for assignment_group_assignment in assignment_group['assignments']:
                if assignment_group_assignment['id'] == assignment['id']:
                    ws[col_weight+str(i)] = assignment_group['group_weight']/len(assignment_group['assignments'])
        if assignment['muted'] == False:
            ws[col_muted+str(i)] = 'No'
        else:
            ws[col_muted+str(i)] = 'Yes'
        
        # Colour the relevant lines
        if ws[col_asgntype+str(i)].value == 'Paper':
            for cell in ws[str(i)+':'+str(i)]:
                cell.font = Font(color='FF0000')
        if assignment['due_at'] != None:
            if datetime.datetime.strptime(assignment['due_at'], "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(days=14) < today and assignment['needs_grading_count'] > 0:
                for cell in ws[str(i)+':'+str(i)]:
                    cell.fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
            if datetime.datetime.strptime(assignment['due_at'], "%Y-%m-%dT%H:%M:%SZ") < today and assignment['needs_grading_count'] == 0 and assignment['muted'] == False and ws[col_asgntype+str(i)].value != 'Paper':
                for cell in ws[str(i)+':'+str(i)]:
                    cell.font = Font(color='00FF00')
        if course['workflow_state'] != 'available' or not assignment['published']:
            for cell in ws[str(i)+':'+str(i)]:
                cell.font = Font(color='00D3D3D3')
        
        i += 1

# Make the workbook into a filtered table.
ws.auto_filter.ref = 'A1:AA'+str(i)



###############################################################################
# Save the created Excel file with customised filename
#
#filename = name_extenstion+'_'+str(datetime.date.today())
filename = name_extension
file_extension = '.xlsx'
wb.save(filename = output_dir+filename+file_extension)
#print 'Workbook: ' + filename + ' saved'

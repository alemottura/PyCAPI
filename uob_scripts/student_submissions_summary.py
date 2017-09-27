#
#	student_summary.py
#
#	This script reads an Excel file of four columns:
#		student_name, student_login_id, tutor_name, tutor_login_id
#	and loops through every student on that list. While looping through every
#	student in the Excel file, it obtains further information about each student
#	such as the courses the student he is registered on and his submissions to
#	assigments on those courses, and saves that information in a list of dictionaries
#	called 'students'.
#
#	Information obtained currently includes:
#		- Canvas ID
#		- University ID
#		- Last 10 page views on Canvas
#		- Courses the student is registered on
#		- Assignments and submissions in those courses
#
#	The script than emails each tutor an Excel file that contains worksheets
#	for every tutee they are responsible for, summarising the information.
#
#	Finally, the script saves an Excel file that contains a filterable table
#	of all submissions by all students in the original Excel file, highligthing
#	late submissions, missing submissions, submissions handled on paper and
#	graded submissions. This is a handy tool to monitor the progress of all
#	students on a programme.
#
#	Things that need to be set:
#
#	student_list - path to Excel file containing details of students
student_list = '/mnt/metadmin/CANVASBOTS/UG/Input/Tutorial_Groups.xlsx'
#
#	output_dir - path to directory where any files should be saved
output_dir = '/mnt/metadmin/CANVASBOTS/UG/Student_Submissions/'
#
#	ug_canvas_accounts - list of Canvas accounts you manage
#	This is required because if students are registered to courses from other
#	accounts, this script will attempt to obtain information from those courses
#	and fail.
ug_canvas_accounts = [115, 116, 117, 118]
#
#	NOTE: this script sends email using the class defined in
#	PyCAPI/uob_scripts/uob_utils.py
#	If you wish to use the script, you need to check that you set up email
#	appropriately.
#
#
#
#
import PyCAPI
import json
from datetime import datetime
from datetime import date
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import uob_utils




capi = PyCAPI.CanvasAPI()
mail = uob_utils.MailAPI()
now = datetime.now()
today = date.today()





###############################################################################
# Only continue with script if it is term time
#
if uob_utils.TermWeek(today)[0] == 0:
	exit


###############################################################################
# Only continue if today is a week day
#
if today.weekday() > 4:
	exit




###############################################################################
# Obtain details of tutees from Excel spreadsheet, and sort them by tutor id
#
try:
	wb = load_workbook(student_list , read_only = True, data_only = True)
	ws = wb['Sheet1']
except:
	raise RuntimeError('Could not open Excel file containing tutees.')
students = []
i = 2
while ws['A'+str(i)].value != None:
    students.append({'name':ws['A'+str(i)].value, 'id':ws['B'+str(i)].value, 'year':ws['C'+str(i)].value, 'tutor_name':ws['D'+str(i)].value, 'tutor_id':ws['E'+str(i)].value})
    i = i + 1
# Sort the list by tutor_id
students = sorted(students, key=lambda k: "%s %s" % (k['tutor_id'], k['id']))
















###############################################################################
# Get all submissions from active and available courses each student is registered to
#
page_views_payload = {}
courses_payload={}
courses_payload["enrollment_state"] = "active" # Only return active courses
courses_payload["state[]"] = "available" # Only return available courses

# Loop through all students
for student in students:
	
	# Create lists to store submissions for each student
	student['submissions'] = []
	student['missed_submissions'] = []
	student['late_submissions'] = []
	student['graded_submissions'] = []
	
	# Obtain the student profile
	student_details = capi.get("/users/%s/profile" % ('sis_login_id:'+student['id']), single=True)
	student['canvas_id'] = student_details['id']
	
	# Request page views
	page_views = capi.get("/users/%s/page_views" % ('sis_login_id:'+student['id']), payload=page_views_payload, first_page_only=True)
	
	# Store information about the last page viewed
	if len(page_views) == 0:
		print '  Student has not viewed any pages in the requested period'
		student['last_access'] = None
	else:
		student['last_access'] = {}
		student['last_access']['date'] = datetime.strptime(page_views[0]['created_at'], "%Y-%m-%dT%H:%M:%SZ")
		student['last_access']['elapsed_time'] = now - student['last_access']['date']
		student['last_access']['url'] = page_views[0]['url']
		student['last_access']['ip'] = page_views[0]['remote_ip']
	
	# Request list of courses for student
	courses = capi.get("/users/%s/courses" % ('sis_login_id:'+student['id']), payload=courses_payload)
	
	# Loop through all courses
	for course in courses:
		
		# Only proceed if authorised user has rights to access the course
		if course["account_id"] in ug_canvas_accounts:
			
			# Obtain submissions of student in the course
			submissions_payload={}
			submissions_payload["student_ids[]"] = 'sis_login_id:'+student['id']
			submissions_payload["grouped"] = False
			submissions_payload["include[]"] = "assignment"
			submissions = capi.get("/courses/%s/students/submissions" % course["id"], single=True, payload=submissions_payload)
			
			# Loop through all submissions and add them to student information
			for submission in submissions[0]["submissions"]:
				submission['assignment']['course_name'] = course['name']
			student['submissions'].extend(submissions[0]["submissions"])
			
			# If student info has no uni_id, find it using this course
			if 'uni_id' not in student:
				uid_payload={}
				uid_payload["search_term"] = student['canvas_id']
				uid = capi.get("/courses/%s/search_users" % course["id"], payload=uid_payload)
				student['uni_id'] = int(uid[0]['sis_user_id'])
	
	# Loop through all submissions and figure out missing/late/graded submissions
	for submission in student['submissions']:
		if submission['assignment']['submission_types'][0] == 'online_upload' or submission['assignment']['submission_types'][0] == 'online_quiz':
			if submission['missing'] == True:
				student['missed_submissions'].append(submission)
		if submission['late'] == True:
			student['late_submissions'].append(submission)
		if submission['workflow_state'] == 'graded':
			student['graded_submissions'].append(submission)






###############################################################################
# Make and email workbooks for tutors
#
wb = Workbook() # Create a workbook
wb.remove_sheet(wb.active) # Remove the active sheet
current_tutor_id = students[0]['tutor_id'] # Save the first tutor_id so you can check when it changes
current_tutor_name = students[0]['tutor_name'] # Save the first tutor_name so you can use it to send email

# Loop through every student
for student in students:
	
	# if tutor has changed, email summary to tutor and start workbook for new tutor
	if student['tutor_id'] != current_tutor_id:
		wb.save(filename=output_dir+current_tutor_id+'.xlsx') # save workbook
		msg = uob_utils.EMailMessage(current_tutor_name +' <'+ current_tutor_id + '@adf.bham.ac.uk>', 'Tutees Weekly Summary')
		msg.body('Hi,\n\nThis is the weekly summary of progress for your tutees on Canvas.\n\nBest Regards,\nMet&Mat UG Office\n\n')
		msg.attach_file(output_dir+current_tutor_id+'.xlsx') # attach workbook
		if today.weekday() == 0: # only send email if it is a Monday...
			mail.send(current_tutor_id + '@adf.bham.ac.uk', msg) # send email
		wb = Workbook() # Create a new workbook
		wb.remove_sheet(wb.active) # Remove the active sheet
		current_tutor_id = student['tutor_id'] # Store the current tutor_id so you can check when it changes
		current_tutor_name = student['tutor_name'] # Store the current tutor_name so you can use it to send the email
	
	# Create a worksheet for the student and start filling it with information
	ws = wb.create_sheet(title=student['name'])
	ws['A1'] = student['name']
	ws['A1'].font = Font(size=14, bold=True)
	ws.merge_cells('A1:H1')
	i = 3
	
	# If there are missed submissions, add them to the worksheet
	if len(student['missed_submissions']) != 0:
		ws['B'+str(i)] = 'Missed assignments'
		ws['B'+str(i)].font = Font(bold=True)
		ws.merge_cells('B'+str(i)+':C'+str(i))
		i += 1
		for submission in student['missed_submissions']:
			ws['B'+str(i)] = submission['assignment']['name']
			ws.merge_cells('B'+str(i)+':H'+str(i))
			i += 1
		i += 1
	
	# If there are late submissions, add them to the worksheet
	if len(student['late_submissions']) != 0:
		ws['B'+str(i)] = 'Late assignments'
		ws['B'+str(i)].font = Font(bold=True)
		ws['H'+str(i)] = 'Days late'
		ws['H'+str(i)].font = Font(bold=True)
		ws.merge_cells('B'+str(i)+':C'+str(i))
		i += 1
		for submission in student['late_submissions']:
			ws['B'+str(i)] = submission['assignment']['name']
			ws.merge_cells('B'+str(i)+':G'+str(i))
			ws['H'+str(i)] = round(submission['duration_late']/86400,1)
			i += 1
		i += 1
	
	# If there are graded submissions, add them to the worksheet
	if len(student['graded_submissions']) != 0:
		ws['B'+str(i)] = 'Graded assignments'
		ws['B'+str(i)].font = Font(bold=True)
		ws['H'+str(i)] = 'Grade (%)'
		ws['H'+str(i)].font = Font(bold=True)
		ws.merge_cells('B'+str(i)+':C'+str(i))
		i += 1
		for submission in student['graded_submissions']:
			ws['B'+str(i)] = submission['assignment']['name']
			ws.merge_cells('B'+str(i)+':G'+str(i))
			ws['H'+str(i)] = round((float(submission['score'])/float(submission['assignment']['points_possible']))*100, 0)
			i += 1
		i += 1

# If loop is done and the last student has been reached, save workbook and email it to the tutor
wb.save(filename=output_dir+current_tutor_id+'.xlsx') # save workbook
msg = uob_utils.EMailMessage(current_tutor_name +' <'+ current_tutor_id + '@adf.bham.ac.uk>', 'Tutees Weekly Summary')
msg.body('Hi,\n\nThis is the weekly summary of progress for your tutees on Canvas.\n\nBest Regards,\nMet&Mat UG Office\n\n')
msg.attach_file(output_dir+current_tutor_id+'.xlsx') # attach workbook
if today.weekday() == 0: # only send email if it is a Monday...
	mail.send(current_tutor_id + '@adf.bham.ac.uk', msg) # send email











###############################################################################
# Make workbook for UG office
#
wb = Workbook() # create workbook
ws = wb.active # get active sheet

# set columns for data (this way one needs to do this once)
col_uid = 'A'
col_lid = 'B'
col_cid = 'C'
col_nm = 'D'
col_yy = 'E'
col_co = 'F'
col_coid = 'G'
col_asgn = 'H'
col_asgnid = 'I'
col_due = 'J'
col_method = 'K'
col_date = 'L'
col_late = 'M'
col_lated = 'N'
col_graded = 'O'
col_grade = 'P'

# set column widths for columns that re
ws.column_dimensions[col_nm].width = 20
ws.column_dimensions[col_co].width = 30
ws.column_dimensions[col_asgn].width = 30
ws.column_dimensions[col_due].width = 20
ws.column_dimensions[col_date].width = 20

# set headers for workbook
ws[col_uid+'1'] = 'ID'
ws[col_lid+'1'] = 'Login ID'
ws[col_cid+'1'] = 'Canvas ID'
ws[col_nm+'1'] = 'Name'
ws[col_yy+'1'] = 'Year'
ws[col_co+'1'] = 'Course'
ws[col_coid+'1'] = 'Course ID'
ws[col_asgn+'1'] = 'Assignment'
ws[col_asgnid+'1'] = 'Assignment ID'
ws[col_due+'1'] = 'Due on'
ws[col_method+'1'] = 'Online/Paper'
ws[col_date+'1'] = 'Submitted on'
ws[col_late+'1'] = 'Late/Missing'
ws[col_lated+'1'] = 'Days late'
ws[col_graded+'1'] = 'Graded'
ws[col_grade+'1'] = 'Grade'

# Loop through all students
i=2
for student in students:
	
	# Loop through all submissions
	for submission in student['submissions']:
		
		# Write standard info about the submission to the workbook
		ws[col_nm+str(i)] = student['name']
		ws[col_cid+str(i)] = student['canvas_id']
		ws[col_lid+str(i)] = student['id']
		ws[col_uid+str(i)] = student['uni_id']
		ws[col_co+str(i)] = submission['assignment']['course_name']
		ws[col_coid+str(i)] = submission['assignment']['course_id']
		ws[col_asgn+str(i)] = submission['assignment']['name']
		ws[col_asgnid+str(i)] = submission['assignment']['id']
		if submission['assignment']['due_at']:
			ws[col_due+str(i)] = datetime.strptime(submission['assignment']['due_at'], "%Y-%m-%dT%H:%M:%SZ")
		if submission['submitted_at']:
			ws[col_date+str(i)] = datetime.strptime(submission['submitted_at'], "%Y-%m-%dT%H:%M:%SZ")

		# Determine whether it was an online submission
		if submission['assignment']['submission_types'][0] == 'online_upload' or submission['assignment']['submission_types'][0] == 'online_quiz':
			ws[col_method+str(i)] = 'Online'
			if submission['missing'] == True:
				ws[col_late+str(i)] = 'Missing'
				for cell in ws[str(i)+':'+str(i)]:
					cell.font = Font(color='00FF9900')
		else:
			ws[col_method+str(i)] = 'Paper'
			for cell in ws[str(i)+':'+str(i)]:
				cell.font = Font(color='00D3D3D3')
	
		# Determine whether the submission was graded
		if submission['workflow_state'] == 'graded':
			ws[col_graded+str(i)] = 'Yes'
			ws[col_grade+str(i)] = round((float(submission['score'])/float(submission['assignment']['points_possible']))*100, 0)
			for cell in ws[str(i)+':'+str(i)]:
					cell.font = Font(color='00006600')

		# Determine whether submission was late		
		if submission['late'] == True:
			ws[col_late+str(i)] = 'Late'
			ws[col_lated+str(i)] = round(submission['duration_late']/86400,1)
			for cell in ws[str(i)+':'+str(i)]:
				cell.font = Font(color='00FF0000')
	
		i += 1	

# Make the workbook into a filtered table.
ws.auto_filter.ref = 'A1:AA'+str(i)

# Save workbook
wb.save(filename=output_dir+'student_submissions.xlsx')	






# BYE BYE!






#       
#	find_login_id.py
#
#	The purpose of this script is to find the login id for every student in
#	a list of university ids (sis_user_id).
#
#	This script reads an Excel file, assuming column A contains the university
#	ids (sis_user_id), uses this data to find out the login_id and saves that
#	column B.
#
#	Things that need to be set:
#
#	excel_file - the path to the Excel file containing the list of IDs
excel_file = './StudentListIncomplete.xlsx'
#
#
#
import PyCAPI
import json
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color


capi = PyCAPI.CanvasAPI()


try:
	wb = load_workbook(excel_file)
	ws = wb['Sheet1']
except:
	raise RuntimeError('Could not open Excel file containing tutees.')
students = []
i = 2
while ws['A'+str(i)].value != None:
	id_number = ws['A'+str(i)].value
	try:
		student = capi.get("/users/%s/profile" % ('sis_user_id:'+str(id_number)), single=True)
		login_id = student['login_id']
		print str(id_number), login_id
		ws['B'+str(i)] = login_id
	except:
		print str(id_number), 'could not find student'
	i += 1

wb.save(filename=excel_file)






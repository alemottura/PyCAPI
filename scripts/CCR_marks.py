#!/opt/local/bin/python

# This script lists all courses in a Canvas account, and published/unpublished assignments within each

import PyCAPI
import os
import glob
import json
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color


capi = PyCAPI.CanvasAPI()


student_list = '/Users/alessandromottura/Dropbox/Teaching/LC_CCR/Challenges/Answers/ProblemSolvingGrades.xlsx'

os.chdir('/Users/alessandromottura/Dropbox/Teaching/LC_CCR/Challenges/Answers/')

try:
	wb = load_workbook(student_list , read_only = True, data_only = True)
	ws = wb['Sheet1']
except:
	raise RuntimeError('Could not open Excel file containing tutees.')
students = []
i = 2
while ws['A'+str(i)].value != None:
	print ws['A'+str(i)].value, ' ', ws['D'+str(i)].value
	print '  Setting grade: ' + str(ws['K'+str(i)].value)
	#capi.grade_assignment_submission(33531, 142110, ws['B'+str(i)].value, ws['K'+str(i)].value)
	for filename in glob.glob('*'+ws['D'+str(i)].value+'*'):
		print '  Uploading file: ' + filename
		capi.comment_assignment_submission_with_upload(33531, 142110, ws['B'+str(i)].value, filename)
	i = i + 1

	
	
	

#with open('./MED_marks.csv', 'rU') as csvfile:
#	content = csv.reader(csvfile, delimiter=',', quotechar='"')
#	for student in content:
#		user_id = student[1]
#		grade = student[6]
#		
#		comment = "I have noticed that the mark for the presentation is not included here.\n"
#		comment += "Report: " + student[4] + "/100" + "\n"
#		comment += "Presentation: " + student[5] + "/100" + "\n"
#		comment += "Total: " + student[6] + "/100" + "\n"
#		comment += "The presentation is weighted at 30% of the total, the report is 70%.\n"
#
#
#		print student[0]
#		print user_id
##		print grade
#		print comment
#		print ''
#
		#capi.comment_assignment_submission(28638, 104449, user_id, comment)
#		capi.grade_assignment_submission(26914, 96088, user_id, grade, comment)

